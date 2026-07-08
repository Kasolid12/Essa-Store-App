"""
Excel Importer — Import data from Excel files into the database.

Functions:
  - import_sku_from_excel(filepath, session) -> dict
  - import_tarif_from_excel(filepath, session) -> dict

Expected Excel formats match the existing template files in the project root
(MasterSKU.xlsx, Master_tarif.xlsx).
"""

import openpyxl
import re
from data.models.sku import SkuMaster
from data.models.master import TarifMaster
from data.models.salary import MasterTarifPenjahit


# ---------------------------------------------------------------------------
#  Import Master SKU
# ---------------------------------------------------------------------------

SKU_EXPECTED_HEADERS = ["Nomor SKU", "Judul", "Rata-Rata Modal Bobot"]


def import_sku_from_excel(filepath: str, session) -> dict:
    """Import Master SKU from Excel (MasterSKU.xlsx format).

    Sheet names tried: "Master SKU" → "Sheet2".
    Upserts by ``kode_sku`` — existing rows are updated.

    Returns
    -------
    dict
        ``{"imported": int, "skipped": int, "errors": [str, ...]}``
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Pick the first matching sheet
    sheet = None
    for candidate in ("Master SKU", "Sheet2"):
        if candidate in wb.sheetnames:
            sheet = wb[candidate]
            break
    if sheet is None:
        raise ValueError(
            f"Tidak menemukan sheet 'Master SKU' atau 'Sheet2' dalam file. "
            f"Sheet tersedia: {wb.sheetnames}"
        )

    # Validate header row
    header = [str(c.value or "").strip() for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    # Normalise: ignore empty trailing columns
    header = [h for h in header if h]
    if header != SKU_EXPECTED_HEADERS:
        raise ValueError(
            f"Header sheet tidak sesuai.\n"
            f"Ekspektasi: {SKU_EXPECTED_HEADERS}\n"
            f"Ditemukan: {header}"
        )

    stats = {"imported": 0, "skipped": 0, "errors": []}

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            kode_sku = _clean_str(row[0]) if row[0] else ""
            nama_produk = _clean_str(row[1]) if row[1] else ""
            modal_raw = str(row[2] or "0").strip().replace(",", "").replace(".", "")
            harga_modal = _parse_float(modal_raw)

            if not kode_sku:
                stats["skipped"] += 1
                continue
            if not nama_produk:
                nama_produk = kode_sku  # fallback

            # Upsert by kode_sku
            existing = (
                session.query(SkuMaster)
                .filter(SkuMaster.kode_sku == kode_sku)
                .first()
            )
            if existing:
                existing.nama_produk = nama_produk
                existing.harga_modal = harga_modal
                # Re-activate if previously soft-deleted
                existing.is_deleted = 0
                existing.is_active = 1
            else:
                session.add(
                    SkuMaster(
                        kode_sku=kode_sku,
                        nama_produk=nama_produk,
                        harga_modal=harga_modal,
                    )
                )
            stats["imported"] += 1

        except Exception as e:
            stats["errors"].append(f"Baris {row_idx}: {e}")

    session.commit()
    wb.close()
    return stats


# ---------------------------------------------------------------------------
#  Import Master Tarif
# ---------------------------------------------------------------------------

TARIF_PENGSUP_HEADERS = ["Nomor SKU", "Kain", "Potongan"]
TARIF_PENJAHIT_HEADERS = ["SKU", "Harga Satuan"]


def import_tarif_from_excel(filepath: str, session) -> dict:
    """Import Master Tarif from Excel (Master_tarif.xlsx format).

    Reads two sheets:
      * **SKU_Pengsup** — ``kode_sku``, ``tarif_pengsup_kain``, ``tarif_pengsup_potongan``
      * **SKU_Penjahit** — ``kode_sku``, ``tarif_jahit``

    Upserts by ``kode_sku`` — existing rows are updated (fields merged).

    Returns
    -------
    dict
        ``{"imported": int, "skipped": int, "errors": [str, ...]}``
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    stats = {"imported": 0, "skipped": 0, "errors": []}
    tarif_buffer = {}  # kode_sku -> dict of fields

    # ---- Sheet 1: SKU_Pengsup ----
    if "SKU_Pengsup" in wb.sheetnames:
        _validate_header(wb["SKU_Pengsup"], TARIF_PENGSUP_HEADERS)
        for row_idx, row in enumerate(
            wb["SKU_Pengsup"].iter_rows(min_row=2, values_only=True), start=2
        ):
            try:
                kode_sku = _clean_str(row[0]) if row[0] else ""
                if not kode_sku:
                    continue
                nilai_kain = _parse_float(str(row[1] or "0").strip())
                nilai_potongan = _parse_float(str(row[2] or "0").strip())

                # Skip rows where both Kain and Potongan are empty/zero —
                # these are cost-definition rows (HANGTAG, WOVEN, etc.)
                if nilai_kain == 0 and nilai_potongan == 0:
                    stats["skipped"] += 1
                    continue

                tarif_buffer.setdefault(kode_sku, {})
                if nilai_kain > 0:
                    tarif_buffer[kode_sku]["tarif_pengsup_kain"] = nilai_kain
                if nilai_potongan > 0:
                    tarif_buffer[kode_sku]["tarif_pengsup_potongan"] = nilai_potongan

            except Exception as e:
                stats["errors"].append(f"Sheet SKU_Pengsup, baris {row_idx}: {e}")

    # ---- Sheet 2: SKU_Penjahit ----
    if "SKU_Penjahit" in wb.sheetnames:
        _validate_header(wb["SKU_Penjahit"], TARIF_PENJAHIT_HEADERS)
        for row_idx, row in enumerate(
            wb["SKU_Penjahit"].iter_rows(min_row=2, values_only=True), start=2
        ):
            try:
                kode_sku = _clean_str(row[0]) if row[0] else ""
                if not kode_sku:
                    continue
                tarif_jahit = _parse_float(str(row[1] or "0").strip())

                # Skip rows without a tarif_jahit value
                # (cost-definition rows like PACK, HANGTAG, etc.)
                if tarif_jahit == 0:
                    stats["skipped"] += 1
                    continue

                tarif_buffer.setdefault(kode_sku, {})
                if tarif_jahit > 0:
                    tarif_buffer[kode_sku]["tarif_jahit"] = tarif_jahit

            except Exception as e:
                stats["errors"].append(f"Sheet SKU_Penjahit, baris {row_idx}: {e}")

    # ---- Persist to database ----
    for kode_sku, fields in tarif_buffer.items():
        try:
            existing = (
                session.query(TarifMaster)
                .filter(TarifMaster.kode_sku == kode_sku)
                .first()
            )
            if existing:
                for key, val in fields.items():
                    setattr(existing, key, val)
            else:
                session.add(
                    TarifMaster(
                        kode_sku=kode_sku,
                        tarif_jahit=fields.get("tarif_jahit", 0.0),
                        tarif_pengsup_kain=fields.get("tarif_pengsup_kain", 0.0),
                        tarif_pengsup_potongan=fields.get("tarif_pengsup_potongan", 0.0),
                    )
                )

            # --- SINKRON: jika ada tarif_jahit, upsert juga ke MasterTarifPenjahit ---
            # (agar SKU Penjahit muncul di dropdown Payroll → tab Penjahit)
            tarif_jahit = fields.get("tarif_jahit", 0.0)
            if tarif_jahit > 0:
                existing_penj = (
                    session.query(MasterTarifPenjahit)
                    .filter(MasterTarifPenjahit.kode_garapan == kode_sku)
                    .first()
                )
                if existing_penj:
                    existing_penj.kode_garapan = kode_sku
                    existing_penj.harga = tarif_jahit
                    existing_penj.is_active = 1
                else:
                    session.add(
                        MasterTarifPenjahit(
                            kode_garapan=kode_sku,
                            harga=tarif_jahit,
                            is_active=1,
                        )
                    )

            stats["imported"] += 1
        except Exception as e:
            stats["errors"].append(f"Upsert {kode_sku}: {e}")

    session.commit()
    wb.close()
    return stats


# ---------------------------------------------------------------------------
#  Helpers
# ---------------------------------------------------------------------------

def _clean_str(val: str) -> str:
    return str(val).strip()


def _parse_float(raw: str) -> float:
    """Parse a numeric string that may contain thousand separators or 'Rp' prefix."""
    cleaned = re.sub(r"[^\d.,\-]", "", raw)  # strip non-numeric except .,-
    # Handle "1.234,56" EU-style → replace last . or use ,
    if "," in cleaned and "." in cleaned:
        # If there's both: assume Indonesian format (dots = thousand, comma = decimal)
        cleaned = cleaned.replace(".", "")
        cleaned = cleaned.replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0


def _validate_header(sheet, expected_headers: list):
    """Check that the first row of *sheet* matches *expected_headers* (lenient)."""
    first_row = [str(c.value or "").strip() for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    # Strip empty trailing cells
    first_row = [h for h in first_row if h]
    # Compare: must start with expected (extra cols allowed)
    for i, exp in enumerate(expected_headers):
        if i >= len(first_row) or first_row[i] != exp:
            raise ValueError(
                f"Header sheet '{sheet.title}' tidak sesuai.\n"
                f"Ekspektasi kolom ke-{i+1}: '{exp}'\n"
                f"Ditemukan: {first_row}"
            )
