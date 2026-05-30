import sqlite3
import os
import openpyxl

def run_import():
    db_path = 'essa.db'
    excel_path = 'Master_tarif.xlsx' # Pastikan nama file Excel kamu sudah sesuai

    if not os.path.exists(db_path):
        print(f"❌ Error: Database {db_path} tidak ditemukan!")
        return
        
    if not os.path.exists(excel_path):
        print(f"❌ Error: File Excel '{excel_path}' tidak ditemukan di folder ini!")
        return

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    print("1. Memeriksa/Membuat tabel tarif_master di database...")
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tarif_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kode_sku VARCHAR NOT NULL UNIQUE,
            tarif_jahit FLOAT DEFAULT 0.0,
            tarif_pengsup_kain FLOAT DEFAULT 0.0,
            tarif_pengsup_potongan FLOAT DEFAULT 0.0
        )
    """)

    # Membaca file Excel (data_only=True agar membaca nilai akhirnya, bukan rumusnya)
    try:
        print(f"Membuka file {excel_path}...")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"❌ Gagal membuka file Excel: {e}")
        return

    # ========================================================
    # 2. IMPORT DATA DARI SHEET: SKU_Penjahit
    # ========================================================
    if 'SKU_Penjahit' in wb.sheetnames:
        print("2. Membaca data dari sheet 'SKU_Penjahit'...")
        sheet_penjahit = wb['SKU_Penjahit']
        
        # Mulai dari baris ke-2 (melewati header "SKU" dan "Harga Satuan")
        for row in sheet_penjahit.iter_rows(min_row=2, values_only=True):
            sku = str(row[0]).strip() if row[0] else ""
            harga_str = str(row[1]).replace(',', '').strip() if row[1] else "0"
            
            try:
                harga = float(harga_str)
            except ValueError:
                harga = 0.0

            # Jika SKU tidak kosong, masukkan ke database
            if sku and sku != "None":
                cursor.execute("""
                    INSERT INTO tarif_master (kode_sku, tarif_jahit)
                    VALUES (?, ?)
                    ON CONFLICT(kode_sku) DO UPDATE SET tarif_jahit=excluded.tarif_jahit
                """, (sku, harga))
    else:
        print("⚠️ Peringatan: Sheet dengan nama 'SKU_Penjahit' tidak ditemukan!")

    # ========================================================
    # 3. IMPORT DATA DARI SHEET: SKU_Pengsup
    # ========================================================
    if 'SKU_Pengsup' in wb.sheetnames:
        print("3. Membaca data dari sheet 'SKU_Pengsup'...")
        sheet_pengsup = wb['SKU_Pengsup']
        
        # Mulai dari baris ke-2 (melewati header "Nomor SKU", "Kain", "Potongan")
        for row in sheet_pengsup.iter_rows(min_row=2, values_only=True):
            sku = str(row[0]).strip() if row[0] else ""
            kain_str = str(row[1]).replace(',', '').strip() if row[1] else "0"
            potongan_str = str(row[2]).replace(',', '').strip() if row[2] else "0"
            
            try:
                kain = float(kain_str)
            except ValueError:
                kain = 0.0
                
            try:
                potongan = float(potongan_str)
            except ValueError:
                potongan = 0.0

            if sku and sku != "None":
                cursor.execute("""
                    INSERT INTO tarif_master (kode_sku, tarif_pengsup_kain, tarif_pengsup_potongan)
                    VALUES (?, ?, ?)
                    ON CONFLICT(kode_sku) DO UPDATE SET 
                    tarif_pengsup_kain=excluded.tarif_pengsup_kain,
                    tarif_pengsup_potongan=excluded.tarif_pengsup_potongan
                """, (sku, kain, potongan))
    else:
        print("⚠️ Peringatan: Sheet dengan nama 'SKU_Pengsup' tidak ditemukan!")

    conn.commit()
    conn.close()
    print("✅ Selesai! Data Master Tarif sukses diimpor dan disinkronisasi dari file Excel.")

if __name__ == '__main__':
    run_import()