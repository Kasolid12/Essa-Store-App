# app_essa/ui/views/gaji_view.py
import os
import pandas as pd
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame, 
    QComboBox, QSpinBox, QDoubleSpinBox, QDateEdit, QMessageBox, 
    QTableWidgetItem, QHeaderView, QTabWidget, QGridLayout, QLineEdit, QGroupBox,
    QFileDialog
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor
from PySide6.QtWidgets import QAbstractItemView, QApplication
from ui.components.tables import CyberTable
from ui.components.buttons import CyberButton
from ui.theme import Theme
from data.database import SessionLocal
from data.models import Person, SkuMaster
from data.models.salary import SalaryRun, SalaryLineItem, MasterTarifPenjahit, PengsupReconciliation
from data.models.master import TarifMaster
from data.models.bon import BonBalance, BonMovement
from utils.pdf_engine import generate_salary_slip
from PySide6.QtCore import QThread, Signal
import os

class PDFWorker(QThread):
    # Sinyal untuk berkomunikasi dengan Main Thread (GajiView)
    finished = Signal(int, str)  # Mengirim (jumlah_sukses, path_folder_export)
    error = Signal(str)          # Mengirim pesan error
    progress = Signal(int, int)  # (Opsional) Mengirim (slip_ke, total_slip)

    def __init__(self, run_ids):
        super().__init__()
        self.run_ids = run_ids

    def run(self):
        """Metode ini berjalan di thread terpisah (Tidak akan bikin UI Not Responding)"""
        try:
            from utils.pdf_engine import generate_salary_slip
            
            export_dir = ""
            total_ids = len(self.run_ids)

            for i, s_id in enumerate(self.run_ids):
                pdf_path = generate_salary_slip(s_id)
                
                if not export_dir and pdf_path:
                    export_dir = os.path.dirname(pdf_path)
                self.progress.emit(i + 1, total_ids)

            self.finished.emit(total_ids, export_dir)
            
        except Exception as e:
            self.error.emit(str(e))

class GajiView(QWidget):
    def __init__(self, notifier=None):
        super().__init__()
        self.db = SessionLocal()
        self.notifier = notifier # Simpan referensi notifier global

        self.cart_penjahit = []
        self.cart_pengsup = []

        self.setup_ui()
        self.load_dropdowns()
        self.load_karyawan_data()

        if self.notifier:
            self.notifier.database_changed.connect(self.refresh_all_gaji_data)

    def refresh_all_gaji_data(self):
        """Fungsi pembungkus untuk memuat ulang semua data dari database terbaru"""
        self.db.expire_all() # Bersihkan cache ORM SQLAlchemy agar riil membaca sqlite terbaru
        self.load_dropdowns()
        self.load_karyawan_data()
        if hasattr(self, 'refresh_pengsup_table'): self.refresh_pengsup_table()
        if hasattr(self, 'recalc_pasukan'): self.recalc_pasukan()

    def parse_waktu(self, v):
        """Helper untuk membaca waktu aneh dari Excel mesin fingerprint"""
        v_str = str(v).strip()
        if v_str in ['', 'nan', 'None']: return None
        if ":" in v_str:
            try: return (int(v_str.split(":")[0]) * 60 + int(v_str.split(":")[1])) / 1440.0
            except: return None
        try:
            if 0.0 <= float(v_str) <= 1.0: return float(v_str)
        except: pass
        return None
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        
        header_layout = QHBoxLayout()
        title = QLabel("PAYROLL & REKAP GAJI")
        title.setStyleSheet(f"font-size: 24pt; font-weight: bold; color: {Theme.NEON_CYAN};")
        header_layout.addWidget(title)
        header_layout.addStretch()
        layout.addLayout(header_layout)

        self.tabs = QTabWidget()
        self.tabs.setStyleSheet(f"""
            QTabBar::tab {{
                background: {Theme.BG_VOID}; color: {Theme.TEXT_MUTED};
                border: 1px solid {Theme.BORDER_DIM}; padding: 12px 25px; font-weight: bold; font-size: 11pt;
            }}
            QTabBar::tab:selected {{
                background: {Theme.BG_PANEL}; color: {Theme.NEON_CYAN};
                border-bottom: 2px solid {Theme.NEON_CYAN};
            }}
            QTabWidget::pane {{ border: 1px solid {Theme.BORDER_DIM}; top: -1px; }}
        """)

        self.setup_tab_penjahit()
        self.setup_tab_pengsup()
        self.setup_tab_pasukan()
        self.setup_tab_bon()

        self.tabs.addTab(self.tab_penjahit, "GAJI PENJAHIT")
        self.tabs.addTab(self.tab_pengsup, "TOTALAN PENGSUP")
        self.tabs.addTab(self.tab_pasukan, "GAJI KARYAWAN (ABSENSI)")
        self.tabs.addTab(self.tab_bon, "KASBON / UTANG")
        
        layout.addWidget(self.tabs)

    # ==========================================
    # TAB 1: GAJI PENJAHIT
    # ==========================================
    def setup_tab_penjahit(self):
        self.tab_penjahit = QWidget()
        lay = QVBoxLayout(self.tab_penjahit)
        
        top_frame = QFrame()
        top_frame.setObjectName("GridPanel")
        top_lay = QGridLayout(top_frame)
        
        self.penj_person = QComboBox()
        self.penj_person.currentIndexChanged.connect(self.on_penjahit_selected)
        self.penj_date = QDateEdit(QDate.currentDate())
        self.penj_date.setCalendarPopup(True)
        
        self.penj_bon_lama = QDoubleSpinBox()
        self.penj_bon_lama.setRange(0, 999999999); self.penj_bon_lama.setPrefix("Rp "); self.penj_bon_lama.setReadOnly(True)
        self.penj_bon_lama.setStyleSheet(f"color: {Theme.NEON_PINK}; font-weight: bold;")
        self.penj_bon_lama.buttonSymbols = QDoubleSpinBox.ButtonSymbols.NoButtons
        
        self.penj_tambah_bon = QDoubleSpinBox()
        self.penj_tambah_bon.setRange(0, 999999999); self.penj_tambah_bon.setPrefix("Rp ")
        
        self.penj_potong_bon = QDoubleSpinBox()
        self.penj_potong_bon.setRange(0, 999999999); self.penj_potong_bon.setPrefix("Rp ")
        
        top_lay.addWidget(QLabel("Nama Penjahit:"), 0, 0); top_lay.addWidget(self.penj_person, 0, 1)
        top_lay.addWidget(QLabel("Tanggal:"), 0, 2); top_lay.addWidget(self.penj_date, 0, 3)
        
        top_lay.addWidget(QLabel("Sisa Bon Lama:"), 1, 0); top_lay.addWidget(self.penj_bon_lama, 1, 1)
        top_lay.addWidget(QLabel("+ Tambah Bon Baru:"), 1, 2); top_lay.addWidget(self.penj_tambah_bon, 1, 3)
        top_lay.addWidget(QLabel("- Potong Bon Minggu Ini:"), 1, 4); top_lay.addWidget(self.penj_potong_bon, 1, 5)
        
        lay.addWidget(top_frame)
        
        mid_frame = QFrame()
        mid_frame.setStyleSheet(f"background-color: {Theme.BG_PANEL}; border: 1px solid {Theme.BORDER_DIM};")
        mid_lay = QHBoxLayout(mid_frame)
        
        self.penj_sku = QComboBox()
        self.penj_sku.setEditable(True)
        self.penj_sku.setMinimumWidth(250)
        self.penj_sku.currentIndexChanged.connect(self.on_garapan_changed)
        
        self.penj_qty = QSpinBox()
        self.penj_qty.setRange(1, 99999)
        self.penj_qty.setPrefix("Qty: ")
        
        self.penj_harga = QDoubleSpinBox()
        self.penj_harga.setRange(0, 999999999); self.penj_harga.setPrefix("Rp ")
        self.penj_harga.setReadOnly(True)
        self.penj_harga.setStyleSheet(f"background-color: {Theme.BG_VOID}; color: {Theme.NEON_YELLOW};")
        self.penj_harga.buttonSymbols = QDoubleSpinBox.ButtonSymbols.NoButtons
        
        btn_add = CyberButton("TAMBAH GARAPAN")
        btn_add.clicked.connect(self.add_garapan)
        
        btn_import_penj = CyberButton("IMPORT EXCEL")
        btn_import_penj.setStyleSheet(f"background-color: {Theme.NEON_PINK}; color: #000; font-weight: bold;")
        btn_import_penj.clicked.connect(self.import_excel_penjahit)
        
        btn_export_penj = CyberButton("EXPORT FORMAT")
        btn_export_penj.setStyleSheet(f"background-color: {Theme.NEON_CYAN}; color: #000; font-weight: bold;")
        btn_export_penj.clicked.connect(self.export_excel_penjahit)
        
        mid_lay.addWidget(QLabel("Pilih SKU/Jenis:")); mid_lay.addWidget(self.penj_sku)
        mid_lay.addWidget(self.penj_qty); mid_lay.addWidget(QLabel("Harga/Pcs:")); mid_lay.addWidget(self.penj_harga)
        mid_lay.addWidget(btn_add)
        mid_lay.addWidget(btn_import_penj)
        mid_lay.addWidget(btn_export_penj)
        
        lay.addWidget(mid_frame)
        
        self.table_penj = CyberTable()
        self.table_penj.setColumnCount(4)
        self.table_penj.setHorizontalHeaderLabels(["Jenis Garapan", "Qty", "Harga Satuan", "Total Harga"])
        self.table_penj.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        lay.addWidget(self.table_penj)
        
        bot_lay = QHBoxLayout()
        btn_del = CyberButton("Hapus Baris", is_danger=True)
        btn_del.clicked.connect(self.del_garapan)
        
        self.lbl_gaji_kotor_penj = QLabel("TOTAL KOTOR: Rp 0")
        self.lbl_gaji_kotor_penj.setStyleSheet(f"font-size: 14pt; font-weight: bold; color: {Theme.NEON_YELLOW};")
        
        btn_save = CyberButton("SIMPAN GAJI PENJAHIT")
        btn_save.setStyleSheet(f"background-color: {Theme.NEON_CYAN}; color: #000; font-weight: bold;")
        btn_save.clicked.connect(self.submit_penjahit)
        
        bot_lay.addWidget(btn_del); bot_lay.addStretch()
        bot_lay.addWidget(self.lbl_gaji_kotor_penj); bot_lay.addWidget(btn_save)
        lay.addLayout(bot_lay)

    # ==========================================
    # TAB 2: TOTALAN PENGSUP
    # ==========================================
    def setup_tab_pengsup(self):
        self.tab_pengsup = QWidget()
        lay = QVBoxLayout(self.tab_pengsup)
        
        top_frame = QFrame()
        top_frame.setObjectName("GridPanel")
        top_lay = QGridLayout(top_frame)
        
        self.psup_person = QComboBox()
        self.psup_person.currentIndexChanged.connect(self.on_pengsup_selected)
        self.psup_date = QDateEdit(QDate.currentDate())
        self.psup_date.setCalendarPopup(True)
        
        self.psup_bon_lama = QDoubleSpinBox()
        self.psup_bon_lama.setRange(0, 999999999); self.psup_bon_lama.setPrefix("Rp "); self.psup_bon_lama.setReadOnly(True)
        self.psup_bon_lama.setStyleSheet(f"color: {Theme.NEON_PINK}; font-weight: bold;")
        self.psup_bon_lama.buttonSymbols = QDoubleSpinBox.ButtonSymbols.NoButtons
        
        self.psup_tambah_bon = QDoubleSpinBox()
        self.psup_tambah_bon.setRange(0, 999999999); self.psup_tambah_bon.setPrefix("Rp ")
        
        self.psup_potong_bon = QDoubleSpinBox()
        self.psup_potong_bon.setRange(0, 999999999); self.psup_potong_bon.setPrefix("Rp ")
        self.psup_potong_bon.valueChanged.connect(self.refresh_pengsup_table)

        self.psup_kain_qty = QDoubleSpinBox()
        self.psup_kain_qty.setRange(0, 99999)
        self.psup_kain_qty.valueChanged.connect(self.refresh_pengsup_table)
        
        self.psup_kain_harga = QDoubleSpinBox()
        self.psup_kain_harga.setRange(0, 999999999); self.psup_kain_harga.setPrefix("Rp ")
        self.psup_kain_harga.valueChanged.connect(self.refresh_pengsup_table)

        top_lay.addWidget(QLabel("Nama Pengsup:"), 0, 0); top_lay.addWidget(self.psup_person, 0, 1)
        top_lay.addWidget(QLabel("Tanggal:"), 0, 2); top_lay.addWidget(self.psup_date, 0, 3)
        
        top_lay.addWidget(QLabel("Sisa Bon Lama:"), 1, 0); top_lay.addWidget(self.psup_bon_lama, 1, 1)
        top_lay.addWidget(QLabel("+ Tambah Bon:"), 1, 2); top_lay.addWidget(self.psup_tambah_bon, 1, 3)
        top_lay.addWidget(QLabel("- Potong Kasbon:"), 1, 4); top_lay.addWidget(self.psup_potong_bon, 1, 5)

        line = QFrame(); line.setFrameShape(QFrame.Shape.HLine); line.setStyleSheet("background-color: #2d2d38;")
        top_lay.addWidget(line, 2, 0, 1, 6)

        top_lay.addWidget(QLabel("PENGURANG (RAW MATERIAL):"), 3, 0)
        top_lay.addWidget(QLabel("Qty Kain Mentah:"), 3, 2); top_lay.addWidget(self.psup_kain_qty, 3, 3)
        top_lay.addWidget(QLabel("Harga Kain/Kg:"), 3, 4); top_lay.addWidget(self.psup_kain_harga, 3, 5)
        
        lay.addWidget(top_frame)

        mid_frame = QFrame()
        mid_frame.setStyleSheet(f"background-color: {Theme.BG_PANEL}; border: 1px solid {Theme.BORDER_DIM};")
        mid_lay = QHBoxLayout(mid_frame)
        
        self.psup_tipe = QComboBox()
        self.psup_tipe.addItems(["Setor Barang Jadi (Kain)", "Setor Potongan (Pcs)", "Potongan Kain Mentah (Kg)"])
        self.psup_tipe.currentIndexChanged.connect(self.on_psup_input_changed)

        self.psup_sku = QComboBox()
        self.psup_sku.setEditable(True)
        self.psup_sku.setMinimumWidth(220)
        self.psup_sku.currentIndexChanged.connect(self.on_psup_input_changed)
        
        self.psup_qty = QDoubleSpinBox()
        self.psup_qty.setRange(0.01, 99999)
        self.psup_qty.setPrefix("Qty: ")
        
        self.psup_harga = QDoubleSpinBox()
        self.psup_harga.setRange(0, 999999999); self.psup_harga.setPrefix("Rp ")
        self.psup_harga.setReadOnly(True)
        self.psup_harga.setStyleSheet(f"background-color: {Theme.BG_VOID}; color: {Theme.NEON_YELLOW};")
        self.psup_harga.buttonSymbols = QDoubleSpinBox.ButtonSymbols.NoButtons
        
        btn_add_psup = CyberButton("TAMBAH KE DAFTAR")
        btn_add_psup.clicked.connect(self.add_pengsup_item)
        
        btn_import_psup = CyberButton("IMPORT EXCEL")
        btn_import_psup.setStyleSheet(f"background-color: {Theme.NEON_PINK}; color: #000; font-weight: bold;")
        btn_import_psup.clicked.connect(self.import_excel_pengsup)
        
        btn_export_psup = CyberButton("EXPORT FORMAT")
        btn_export_psup.setStyleSheet(f"background-color: {Theme.NEON_CYAN}; color: #000; font-weight: bold;")
        btn_export_psup.clicked.connect(self.export_excel_pengsup)

        mid_lay.addWidget(self.psup_tipe)
        mid_lay.addWidget(QLabel("SKU:")); mid_lay.addWidget(self.psup_sku)
        mid_lay.addWidget(self.psup_qty)
        mid_lay.addWidget(QLabel("Harga/Unit:")); mid_lay.addWidget(self.psup_harga)
        mid_lay.addWidget(btn_add_psup)
        
        lay.addWidget(mid_frame)

        self.table_psup = CyberTable()
        self.table_psup.setColumnCount(5)
        self.table_psup.setHorizontalHeaderLabels(["Tipe Setoran", "Kode / Nama SKU", "Qty", "Harga Satuan", "Total Harga"])
        self.table_psup.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        lay.addWidget(self.table_psup)
        
        bot_lay = QHBoxLayout(); bot_lay.setSpacing(20)
        
        left_panel = QGroupBox("CATATAN NOTA & OPERASIONAL TABEL")
        left_panel.setStyleSheet(f"QGroupBox {{ color: {Theme.NEON_CYAN}; font-weight: bold; border: 1px solid #2d2d38; border-radius: 8px; margin-top: 5px; padding: 15px; }}")
        left_lay = QVBoxLayout(left_panel); left_lay.setSpacing(12)
        
        self.psup_catatan_manual = QLineEdit()
        self.psup_catatan_manual.setPlaceholderText("Ketik keterangan tambahan di sini jika ada...")
        self.psup_catatan_manual.setStyleSheet(f"background-color: #15151a; color: {Theme.TEXT_MAIN}; padding: 8px; border: 1px solid #2d2d38; border-radius: 4px;")
        
        btn_del_psup = CyberButton("Hapus Baris Terpilih", is_danger=True)
        btn_del_psup.clicked.connect(self.del_pengsup_item)
        btn_del_psup.setMinimumHeight(40)
        
        left_lay.addWidget(QLabel("Tulis Catatan Manual:")); left_lay.addWidget(self.psup_catatan_manual)
        left_lay.addWidget(btn_del_psup); left_lay.addStretch()
        left_lay.addWidget(btn_import_psup)
        left_lay.addWidget(btn_export_psup)
        
        right_panel = QGroupBox("SUMMARY PERHITUNGAN FINANSIAL")
        right_panel.setStyleSheet(f"QGroupBox {{ color: {Theme.NEON_YELLOW}; font-weight: bold; border: 1px solid #2d2d38; border-radius: 8px; margin-top: 5px; padding: 15px; }}")
        right_lay = QVBoxLayout(right_panel); right_lay.setSpacing(10)
        
        self.lbl_pemasukan = QLabel("Total Pemasukan Barang/Potongan: Rp 0"); self.lbl_pemasukan.setStyleSheet(f"color: {Theme.TEXT_MAIN}; font-size: 10pt;")
        self.lbl_pengurang = QLabel("- Potongan Harga Kain Mentah: Rp 0"); self.lbl_pengurang.setStyleSheet("color: #ff5252; font-size: 10pt;")
        self.lbl_gaji_kotor_psup = QLabel("TOTAL DITERIMA (SBLM KASBON): Rp 0"); self.lbl_gaji_kotor_psup.setStyleSheet(f"font-size: 11pt; font-weight: bold; color: {Theme.NEON_YELLOW}; border-top: 1px solid #2d2d38; padding-top: 5px;")
        self.lbl_gaji_bersih_psup = QLabel("GAJI BERSIH (NETT): Rp 0"); self.lbl_gaji_bersih_psup.setStyleSheet(f"font-size: 13pt; font-weight: bold; color: {Theme.NEON_CYAN};")
        
        btn_save_psup = CyberButton("SIMPAN REKAP PENGSUP")
        btn_save_psup.setStyleSheet(f"background-color: {Theme.NEON_CYAN}; color: #000; font-weight: bold; font-size: 11pt; padding: 10px;")
        btn_save_psup.clicked.connect(self.submit_pengsup)
        
        right_lay.addWidget(self.lbl_pemasukan); right_lay.addWidget(self.lbl_pengurang)
        right_lay.addWidget(self.lbl_gaji_kotor_psup); right_lay.addWidget(self.lbl_gaji_bersih_psup)
        right_lay.addWidget(btn_save_psup)
        
        bot_lay.addWidget(left_panel, stretch=1); bot_lay.addWidget(right_panel, stretch=1)
        lay.addLayout(bot_lay)

    # ==========================================
    # TAB 3: GAJI KARYAWAN (ABSENSI) - NEW SMART GRID
    # ==========================================
    def setup_tab_pasukan(self):
        self.tab_pasukan = QWidget()
        lay = QVBoxLayout(self.tab_pasukan)

        # -- Top Bar Settings --
        top_frame = QFrame()
        top_frame.setObjectName("GridPanel")
        top_lay = QHBoxLayout(top_frame)

        btn_import = CyberButton("📥 IMPORT EXCEL ABSENSI")
        btn_import.setStyleSheet(f"background-color: {Theme.NEON_YELLOW}; color: #000; font-weight: bold;")
        btn_import.clicked.connect(self.import_absensi_excel)
        
        btn_export_pasukan = CyberButton("EXPORT FORMAT")
        btn_export_pasukan.setStyleSheet(f"background-color: {Theme.NEON_CYAN}; color: #000; font-weight: bold;")
        btn_export_pasukan.clicked.connect(self.export_excel_pasukan)
        
        self.lbl_file_absen = QLabel("Input manual di tabel, atau Import Excel")
        self.lbl_file_absen.setStyleSheet(f"color: {Theme.TEXT_MUTED}; font-style: italic;")

        self.pasukan_date = QDateEdit(QDate.currentDate())
        self.pasukan_date.setCalendarPopup(True)

        self.tarif_normal = QDoubleSpinBox()
        self.tarif_normal.setRange(0, 99999); self.tarif_normal.setValue(150)
        self.tarif_normal.setPrefix("Rp "); self.tarif_normal.setSuffix(" /Mnt")
        self.tarif_normal.valueChanged.connect(self.recalc_pasukan)

        self.tarif_lembur = QDoubleSpinBox()
        self.tarif_lembur.setRange(0, 99999); self.tarif_lembur.setValue(160)
        self.tarif_lembur.setPrefix("Rp "); self.tarif_lembur.setSuffix(" /Mnt")
        self.tarif_lembur.valueChanged.connect(self.recalc_pasukan)

        top_lay.addWidget(btn_import)
        top_lay.addWidget(btn_export_pasukan)
        top_lay.addWidget(self.lbl_file_absen)
        top_lay.addStretch()
        top_lay.addWidget(QLabel("Tgl Payroll:")); top_lay.addWidget(self.pasukan_date)
        top_lay.addWidget(QLabel("Tarif Normal:")); top_lay.addWidget(self.tarif_normal)
        top_lay.addWidget(QLabel("Tarif Lembur:")); top_lay.addWidget(self.tarif_lembur)
        lay.addWidget(top_frame)

        # -- Smart Grid Table --
        self.table_pasukan = CyberTable()
        self.table_pasukan.cellChanged.connect(self.on_pasukan_cell_edited)
        self.table_pasukan.setColumnCount(11)
        self.table_pasukan.setHorizontalHeaderLabels([
            "ID", "Nama Karyawan", "Hadir", 
            "Menit Normal", "Tarif Normal",
            "Menit Lembur", "Tarif Lembur",
            "Gaji Kotor", "Bon Lama", "Potong Kasbon", "Gaji Bersih"
        ])
        self.table_pasukan.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked | QAbstractItemView.EditTrigger.AnyKeyPressed)
        # Hubungkan kembali ke fungsi cell edited yang sudah kita amankan
        self.table_pasukan.cellChanged.connect(self.on_pasukan_cell_edited)
        
        self.table_pasukan.hideColumn(0) # Sembunyikan kolom ID database
        self.table_pasukan.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        # Warna background header
        self.table_pasukan.setStyleSheet(f"QHeaderView::section {{ background-color: #1e1e24; color: {Theme.NEON_CYAN}; font-weight: bold; padding: 5px; text-align: center; }}")
        lay.addWidget(self.table_pasukan)

        # -- Bottom Action --
        bot_lay = QHBoxLayout()
        hint = QLabel("💡 TIPS: Klik ganda pada kolom Hadir, Menit, atau Potong Kasbon untuk mengedit nilainya. Lalu klik tombol 'Hitung Ulang'.")
        hint.setStyleSheet(f"color: {Theme.NEON_PINK}; font-style: italic;")
        
        # UBAH KODE TOMBOL REFRESH MENJADI INI:
        btn_reset_pasukan = CyberButton("KOSONGKAN TABEL")
        btn_reset_pasukan.setStyleSheet(f"background-color: {Theme.NEON_PINK}; color: #FFF; font-weight: bold;")
        # Kita panggil fungsi load_karyawan_data karena fungsi itu otomatis me-reset semua baris jadi 0
        btn_reset_pasukan.clicked.connect(self.load_karyawan_data)
        
        btn_save_pasukan = CyberButton("SIMPAN & CETAK SEMUA SLIP")
        btn_save_pasukan.setStyleSheet(f"background-color: {Theme.NEON_CYAN}; color: #000; font-weight: bold; padding: 8px 15px;")
        btn_save_pasukan.clicked.connect(self.submit_pasukan)

        bot_lay.addWidget(hint); bot_lay.addStretch()
        bot_lay.addWidget(btn_reset_pasukan)
        bot_lay.addWidget(btn_save_pasukan)
        lay.addLayout(bot_lay)
    
    # ==========================================
    # 4. TAB MANAJEMEN BON (Telah Di-Upgrade)
    # ==========================================
    def setup_tab_bon(self):
        self.tab_bon = QWidget()
        main_lay = QVBoxLayout(self.tab_bon)
        
        # --- TAB INNER UNTUK BON ---
        self.bon_tabs = QTabWidget()
        self.bon_tabs.setStyleSheet(f"""
            QTabBar::tab {{ background: {Theme.BG_VOID}; color: {Theme.TEXT_MUTED}; border: 1px solid {Theme.BORDER_DIM}; padding: 10px 20px; font-weight: bold; }}
            QTabBar::tab:selected {{ background: {Theme.BG_PANEL}; color: {Theme.NEON_CYAN}; border-bottom: 2px solid {Theme.NEON_CYAN}; }}
            QTabWidget::pane {{ border: 1px solid {Theme.BORDER_DIM}; top: -1px; }}
        """)
        
        # --- 4A. SUB-TAB: DASHBOARD BON (Tampilan Asli) ---
        self.sub_tab_dash_bon = QWidget()
        lay_dash = QVBoxLayout(self.sub_tab_dash_bon)
        
        top_lay = QHBoxLayout()
        top_lay.addWidget(QLabel("Daftar Sisa Kasbon Aktif:"))
        top_lay.addStretch()
        
        self.btn_refresh_bon = CyberButton("🔄 REFRESH DATA BON")
        self.btn_refresh_bon.clicked.connect(self.load_bon)
        top_lay.addWidget(self.btn_refresh_bon)
        lay_dash.addLayout(top_lay)
        
        self.table_bon = CyberTable()
        self.table_bon.setColumnCount(4)
        self.table_bon.setHorizontalHeaderLabels(["ID Person", "Nama Personel", "Tipe Personel", "Sisa Kasbon"])
        self.table_bon.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        lay_dash.addWidget(self.table_bon)
        
        # --- 4B. SUB-TAB: UPDATE BON MANUAL (Baru) ---
        self.sub_tab_manual_bon = QWidget()
        lay_manual = QVBoxLayout(self.sub_tab_manual_bon)
        
        form_group = QGroupBox("FORM UPDATE MANUAL KASBON")
        form_group.setStyleSheet(f"QGroupBox {{ color: {Theme.NEON_YELLOW}; font-weight: bold; border: 1px solid #2d2d38; padding: 15px; margin-top: 10px; }}")
        form_lay = QGridLayout(form_group)
        
        self.cb_bon_person = QComboBox()
        self.cb_bon_person.currentIndexChanged.connect(self.on_bon_manual_person_changed)
        
        self.lbl_bon_saldo = QLabel("Rp 0")
        self.lbl_bon_saldo.setStyleSheet(f"font-size: 16pt; font-weight: bold; color: {Theme.NEON_PINK};")
        
        self.cb_bon_aksi = QComboBox()
        self.cb_bon_aksi.addItems(["+ TAMBAH BON BARU", "- POTONG BON (BAYAR TUNAI)"])
        
        self.spin_bon_nominal = QDoubleSpinBox()
        self.spin_bon_nominal.setRange(0, 999999999)
        self.spin_bon_nominal.setPrefix("Rp ")
        
        self.txt_bon_ket = QLineEdit()
        self.txt_bon_ket.setPlaceholderText("Catatan (Misal: Pinjam darurat, Cicil tunai...)")
        
        btn_submit_manual = CyberButton("SIMPAN UPDATE BON")
        btn_submit_manual.clicked.connect(self.submit_manual_bon)
        
        form_lay.addWidget(QLabel("Pilih Nama:"), 0, 0)
        form_lay.addWidget(self.cb_bon_person, 0, 1)
        form_lay.addWidget(QLabel("Sisa Saldo Saat Ini:"), 0, 2)
        form_lay.addWidget(self.lbl_bon_saldo, 0, 3)
        
        line = QFrame(); line.setFrameShape(QFrame.Shape.HLine); line.setStyleSheet("background-color: #2d2d38;")
        form_lay.addWidget(line, 1, 0, 1, 4)
        
        form_lay.addWidget(QLabel("Jenis Tindakan:"), 2, 0)
        form_lay.addWidget(self.cb_bon_aksi, 2, 1)
        form_lay.addWidget(QLabel("Nominal:"), 2, 2)
        form_lay.addWidget(self.spin_bon_nominal, 2, 3)
        
        form_lay.addWidget(QLabel("Keterangan:"), 3, 0)
        form_lay.addWidget(self.txt_bon_ket, 3, 1, 1, 2)
        form_lay.addWidget(btn_submit_manual, 3, 3)
        
        lay_manual.addWidget(form_group)
        
        lay_manual.addWidget(QLabel("Riwayat Transaksi Kasbon (Orang Terpilih):"))
        self.table_bon_history = CyberTable()
        self.table_bon_history.setColumnCount(5)
        self.table_bon_history.setHorizontalHeaderLabels(["Tanggal", "Jenis", "Sumber", "Keterangan", "Nominal"])
        self.table_bon_history.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        lay_manual.addWidget(self.table_bon_history)
        
        # Satukan ke Tab Utama
        self.bon_tabs.addTab(self.sub_tab_dash_bon, "DASHBOARD KASBON")
        self.bon_tabs.addTab(self.sub_tab_manual_bon, "UPDATE BON MANUAL")
        main_lay.addWidget(self.bon_tabs)

    # ==========================================
    # LOGIC UMUM & PENJAHIT
    # ==========================================
    def load_dropdowns(self):
        self.penj_person.clear(); self.penj_person.addItem("-- Pilih Penjahit --", None)
        self.psup_person.clear(); self.psup_person.addItem("-- Pilih Pengsup --", None)
        # --- PERBAIKAN FINAL DROPDOWN BON ---
        if hasattr(self, 'cb_bon_person'):
            self.cb_bon_person.blockSignals(True)
            self.cb_bon_person.clear()
            self.cb_bon_person.addItem("-- Pilih Nama --", None)
            
            # KODE BARU: Ambil data mandiri langsung dari database agar tidak error
            bon_persons = self.db.query(Person).filter(Person.person_type.in_(['PENJAHIT', 'PENGSUP', 'KARYAWAN'])).order_by(Person.nama).all()
            
            for p in bon_persons:
                self.cb_bon_person.addItem(f"{p.nama} ({p.person_type})", p.id)
                
            self.cb_bon_person.blockSignals(False)
        # ------------------------------------
        
        persons = self.db.query(Person).all()
        for p in persons:
            if p.person_type == 'PENJAHIT': self.penj_person.addItem(p.nama, p.id)
            elif p.person_type == 'PENGSUP': self.psup_person.addItem(p.nama, p.id)
            
        self.penj_sku.clear(); self.penj_sku.addItem("-- Pilih Garapan --", None)
        tarifs_penj = self.db.query(MasterTarifPenjahit).filter(MasterTarifPenjahit.is_active == 1).order_by(MasterTarifPenjahit.kode_garapan).all()
        for t in tarifs_penj: self.penj_sku.addItem(t.kode_garapan, (t.id, t.harga, getattr(t, 'sku_id', None)))
            
        self.psup_sku.clear(); self.psup_sku.addItem("-- Ketik/Pilih SKU --", None)
        
        # Hapus filter is_active dan bungkus dengan str()
        skus = self.db.query(SkuMaster).order_by(SkuMaster.kode_sku).all()
        for s in skus:
            kode_teks = str(s.kode_sku) if s.kode_sku else "NO-KODE"
            self.psup_sku.addItem(kode_teks, kode_teks)
    
    # --- LOGIKA UPDATE BON MANUAL ---
    def on_bon_manual_person_changed(self):
        person_id = self.cb_bon_person.currentData()
        if not person_id:
            self.lbl_bon_saldo.setText("Rp 0")
            self.table_bon_history.setRowCount(0)
            return
            
        balance = self.db.query(BonBalance).filter(BonBalance.person_id == person_id).first()
        saldo_saat_ini = balance.saldo if balance else 0
        self.lbl_bon_saldo.setText(f"Rp {saldo_saat_ini:,.0f}")
        
        movements = self.db.query(BonMovement).filter(BonMovement.person_id == person_id).order_by(BonMovement.id.desc()).limit(30).all()
        self.table_bon_history.setRowCount(len(movements))
        for row, m in enumerate(movements):
            self.table_bon_history.setItem(row, 0, QTableWidgetItem(str(m.tanggal)))
            
            jenis_item = QTableWidgetItem(m.tipe)
            if "TAMBAH" in m.tipe: jenis_item.setForeground(Qt.GlobalColor.yellow)
            else: jenis_item.setForeground(Qt.GlobalColor.green)
            self.table_bon_history.setItem(row, 1, jenis_item)
            
            self.table_bon_history.setItem(row, 2, QTableWidgetItem(str(m.sumber)))
            ket_text = getattr(m, 'keterangan', "-") 
            self.table_bon_history.setItem(row, 3, QTableWidgetItem(str(ket_text)))
            try:
                # Ambil data (coba 'nominal', atau fallback ke 'jumlah')
                raw_val = getattr(m, 'nominal', getattr(m, 'jumlah', 0))
                # Konversi paksa menjadi angka, jika kosong jadikan 0
                val = float(raw_val) if raw_val not in [None, ""] else 0.0
            except Exception:
                val = 0.0
                
            nom_item = QTableWidgetItem(f"Rp {val:,.0f}")
            nom_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table_bon_history.setItem(row, 4, nom_item)
            # -----------------------------------

    def submit_manual_bon(self):
        person_id = self.cb_bon_person.currentData()
        nominal = self.spin_bon_nominal.value()
        is_tambah = (self.cb_bon_aksi.currentIndex() == 0)
        ket = self.txt_bon_ket.text().strip()
        
        if not person_id: return QMessageBox.warning(self, "Error", "Pilih nama terlebih dahulu!")
        if nominal <= 0: return QMessageBox.warning(self, "Error", "Nominal tidak boleh nol!")
            
        try:
            balance = self.db.query(BonBalance).filter(BonBalance.person_id == person_id).first()
            if not balance:
                balance = BonBalance(person_id=person_id, saldo=0)
                self.db.add(balance)
                
            if not is_tambah and nominal > balance.saldo:
                return QMessageBox.warning(self, "Error", "Potongan tidak boleh lebih besar dari sisa saldo!")

            if is_tambah:
                balance.saldo += nominal
                tipe_mov = "TAMBAH_MANUAL"
            else:
                balance.saldo -= nominal
                tipe_mov = "POTONG_MANUAL"
                
            tgl_sekarang = QDate.currentDate().toString("yyyy-MM-dd")
            catatan = ket if ket else "Update Manual Kasir"
            
            movement = BonMovement(
                person_id=person_id, tanggal=tgl_sekarang, tipe=tipe_mov, 
                nominal=nominal, sumber=f"MANUAL: {catatan}"
            )
            self.db.add(movement)
            self.db.commit()
            if hasattr(self, 'notifier') and self.notifier:
                print("[*] Broadcasting database changes to all menus...")
                self.notifier.database_changed.emit()
            QMessageBox.information(self, "Sukses", "Data kasbon berhasil diupdate manual!")
            
            self.spin_bon_nominal.setValue(0)
            self.txt_bon_ket.clear()
            self.on_bon_manual_person_changed()
            if hasattr(self, 'load_bon'): self.load_bon() # Auto-refresh dashboard
                
        except Exception as e:
            self.db.rollback()
            self.db.expire_all()
            QMessageBox.critical(self, "Error", f"Terjadi kesalahan database: {e}")
            
    def load_bon(self):
        self.table_bon.setRowCount(0)
        # Ambil data saldo yang lebih dari 0
        balances = self.db.query(BonBalance).filter(BonBalance.saldo > 0).all()
        
        self.table_bon.setRowCount(len(balances))
        for row, b in enumerate(balances):
            self.table_bon.setItem(row, 0, QTableWidgetItem(str(b.person_id)))
            self.table_bon.setItem(row, 1, QTableWidgetItem(b.person.nama if b.person else "Unknown"))
            self.table_bon.setItem(row, 2, QTableWidgetItem(b.person.person_type if b.person else "-"))
            
            saldo_item = QTableWidgetItem(f"Rp {b.saldo:,.0f}")
            # Warna pink/magenta untuk sisa bon
            saldo_item.setForeground(Qt.GlobalColor.magenta)
            saldo_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table_bon.setItem(row, 3, saldo_item)

    def on_garapan_changed(self, index):
        if index > 0:
            data = self.penj_sku.itemData(index)
            if data: self.penj_harga.setValue(data[1])

    def on_penjahit_selected(self):
        p_id = self.penj_person.currentData()
        if not p_id: self.penj_bon_lama.setValue(0); return
        balance = self.db.query(BonBalance).filter(BonBalance.person_id == p_id).first()
        self.penj_bon_lama.setValue(balance.saldo if balance else 0)

    def add_garapan(self):
        sku_text = self.penj_sku.currentText()
        data = self.penj_sku.currentData()
        
        # Ekstrak data 3-tuple yang baru saja kita modifikasi
        tarif_id = data[0] if data else None
        sku_id = data[2] if data and len(data) > 2 else None
        
        qty, harga = self.penj_qty.value(), self.penj_harga.value()
        if not sku_text or sku_text.startswith("--") or harga <= 0: return
        
        # Masukkan sku_id ke dalam keranjang, bukan None lagi!
        self.cart_penjahit.append({"tarif_id": tarif_id, "sku_id": sku_id, "nama_garapan": sku_text, "qty": qty, "harga": harga, "total": qty * harga})
        
        self.refresh_penjahit_table()
        self.penj_qty.setValue(0)
        
    def del_garapan(self):
        selected = self.table_penj.selectedItems()
        if not selected: return
        rows = sorted(list(set([item.row() for item in selected])), reverse=True)
        for r in rows: self.cart_penjahit.pop(r)
        self.refresh_penjahit_table()

    def refresh_penjahit_table(self):
        self.table_penj.setRowCount(0)
        gaji_kotor = 0
        for i, item in enumerate(self.cart_penjahit):
            self.table_penj.insertRow(i)
            self.table_penj.setItem(i, 0, QTableWidgetItem(item['nama_garapan']))
            self.table_penj.setItem(i, 1, QTableWidgetItem(str(item['qty'])))
            self.table_penj.setItem(i, 2, QTableWidgetItem(f"Rp {item['harga']:,.0f}"))
            self.table_penj.setItem(i, 3, QTableWidgetItem(f"Rp {item['total']:,.0f}"))
            gaji_kotor += item['total']
        self.lbl_gaji_kotor_penj.setText(f"TOTAL KOTOR: Rp {gaji_kotor:,.0f}")

    # ==========================================
    # LOGIC UPDATE PENJAHIT (DB Bug Fix + PDF)
    # ==========================================
    def submit_penjahit(self):
        person_id = self.penj_person.currentData()
        person_nama = self.penj_person.currentText()
        if not person_id or not self.cart_penjahit: return QMessageBox.warning(self, "Error", "Pilih penjahit & pastikan garapan tidak kosong!")

        bon_lama, tambah_bon, potong_bon = self.penj_bon_lama.value(), self.penj_tambah_bon.value(), self.penj_potong_bon.value()
        gaji_kotor = sum(item['total'] for item in self.cart_penjahit)
        gaji_bersih = gaji_kotor - potong_bon
        sisa_bon_akhir = bon_lama + tambah_bon - potong_bon

        if potong_bon > bon_lama: 
            return QMessageBox.warning(self, "Error", "Potongan tidak boleh melebihi Sisa Bon Lama!")
        if potong_bon > gaji_kotor: 
            return QMessageBox.warning(self, "Error", "Potongan bon tidak boleh lebih besar dari Gaji Kotor!")

        try:
            tanggal_str = self.penj_date.date().toString("yyyy-MM-dd")
            run = SalaryRun(tipe="BORONGAN_PENJAHIT", person_id=person_id, tanggal_proses=tanggal_str, gaji_kotor=gaji_kotor,
                            bon_lama=bon_lama, tambah_bon=tambah_bon, potong_bon=potong_bon, gaji_bersih=gaji_bersih, sisa_bon_akhir=sisa_bon_akhir)
            self.db.add(run); self.db.flush()

            items_for_pdf = [] # Kumpulkan untuk PDF

            for item in self.cart_penjahit:
                # FIX DB BUG: Pastikan sku_id di-handle aman jika None
                line = SalaryLineItem(salary_run_id=run.id, sku_id=item.get('sku_id'), model_code=item['nama_garapan'], 
                                      qty=item['qty'], tarif_per_pcs=item['harga'], subtotal=item['total'])
                self.db.add(line)
                items_for_pdf.append({"garapan": item['nama_garapan'], "qty": item['qty'], "harga": item['harga'], "total": item['total']})

            balance = self.db.query(BonBalance).filter(BonBalance.person_id == person_id).first()
            if not balance:
                balance = BonBalance(person_id=person_id, saldo=0)
                self.db.add(balance)

            if tambah_bon > 0:
                self.db.add(BonMovement(person_id=person_id, tanggal=tanggal_str, tipe="TAMBAH", nominal=tambah_bon, sumber="PAYROLL_PENJAHIT"))
                balance.saldo += tambah_bon
            if potong_bon > 0:
                self.db.add(BonMovement(person_id=person_id, tanggal=tanggal_str, tipe="POTONG_GAJI", nominal=potong_bon, sumber="PAYROLL_PENJAHIT"))
                balance.saldo -= potong_bon

            self.db.commit()
            if hasattr(self, 'notifier') and self.notifier:
                print("[*] Broadcasting database changes to all menus...")
                self.notifier.database_changed.emit()
                
            try:
                pdf_path = generate_salary_slip(run.id) # Cukup passing run.id
                if pdf_path and os.path.exists(pdf_path):
                    os.startfile(pdf_path)
            except Exception as pdf_err:
                QMessageBox.warning(self, "PDF Error", f"Data tersimpan, tapi PDF gagal dicetak: {pdf_err}")
            # ------------------------------------

            QMessageBox.information(self, "Sukses", "Data gaji penjahit berhasil disimpan dan PDF dicetak!")
            self.cart_penjahit.clear(); self.refresh_penjahit_table()
            self.penj_tambah_bon.setValue(0); self.penj_potong_bon.setValue(0); self.penj_harga.setValue(0)
            self.on_penjahit_selected()

        except Exception as e:
            self.db.rollback() 
            self.db.expire_all() 
            QMessageBox.critical(self, "Error", f"Gagal Database: {e}")

    def import_excel_penjahit(self):
        """Fitur baru untuk import baris garapan Penjahit secara massal"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Import Garapan Penjahit", "", "Excel Files (*.xlsx *.xls *.csv)")
        if not file_path: return

        try:
            df = pd.read_excel(file_path) if file_path.endswith('.xls') or file_path.endswith('.xlsx') else pd.read_csv(file_path)
            # Standarisasi kolom (Hapus spasi, jadikan huruf kecil)
            df.columns = df.columns.astype(str).str.strip().str.lower()
            
            # Cari kolom yang sesuai
            col_sku = next((c for c in df.columns if 'sku' in c or 'garapan' in c or 'jenis' in c), None)
            col_qty = next((c for c in df.columns if 'qty' in c or 'jumlah' in c), None)
            col_harga = next((c for c in df.columns if 'harga' in c or 'tarif' in c), None)

            if not col_sku or not col_qty:
                return QMessageBox.warning(self, "Format Salah", "Sistem butuh kolom bernama 'SKU/Garapan' dan 'Qty' di file Excel.")

            for _, row in df.iterrows():
                nama = str(row[col_sku]).strip()
                qty = float(row[col_qty])
                
                # Jika harga tidak ada di excel, cari ke database master
                harga = float(row[col_harga]) if col_harga and pd.notnull(row[col_harga]) else 0
                tarif_id = None
                
                if harga == 0:
                    tarif_db = self.db.query(MasterTarifPenjahit).filter(MasterTarifPenjahit.kode_garapan == nama).first()
                    if tarif_db:
                        harga = tarif_db.harga
                        tarif_id = tarif_db.id

                if harga > 0 and qty > 0:
                    self.cart_penjahit.append({"tarif_id": tarif_id, "sku_id": None, "nama_garapan": nama, "qty": qty, "harga": harga, "total": qty * harga})

            self.refresh_penjahit_table()
            QMessageBox.information(self, "Sukses", "Data Excel berhasil ditarik ke keranjang!")
        except Exception as e:
            QMessageBox.critical(self, "Error Import", f"Gagal membaca Excel: {e}")

    # ==========================================
    # LOGIC: PENGSUP
    # ==========================================
    def on_pengsup_selected(self):
        p_id = self.psup_person.currentData()
        if not p_id: self.psup_bon_lama.setValue(0); return
        balance = self.db.query(BonBalance).filter(BonBalance.person_id == p_id).first()
        self.psup_bon_lama.setValue(balance.saldo if balance else 0)

    def get_pengsup_tarif_cerdas(self, sku_kode, is_kain=True):
        tarif = self.db.query(TarifMaster).filter(TarifMaster.kode_sku == sku_kode).first()
        if tarif: return tarif.tarif_pengsup_kain if is_kain else tarif.tarif_pengsup_potongan
            
        parts = str(sku_kode).strip().upper().split('-')
        base = parts[0]
        size = ""
        for p in parts:
            if p in ["S", "M", "L", "XL"]:
                size = p
                break
                
        if size:
            fallback = f"{base}-{size}"
            tarif_f1 = self.db.query(TarifMaster).filter(TarifMaster.kode_sku == fallback).first()
            if tarif_f1: return tarif_f1.tarif_pengsup_kain if is_kain else tarif_f1.tarif_pengsup_potongan
                
        tarif_f2 = self.db.query(TarifMaster).filter(TarifMaster.kode_sku == base).first()
        if tarif_f2: return tarif_f2.tarif_pengsup_kain if is_kain else tarif_f2.tarif_pengsup_potongan
            
        return 0.0

    def on_psup_input_changed(self, index=None):
        sku_kode = self.psup_sku.currentData()
        sku_text = self.psup_sku.currentText()
        if not sku_text or sku_text.startswith("--"): return
        if not sku_kode: sku_kode = sku_text
            
        is_kain = "Kain" in self.psup_tipe.currentText()
        harga = self.get_pengsup_tarif_cerdas(sku_kode, is_kain)
        self.psup_harga.setValue(harga)

    def add_pengsup_item(self):
        tipe = self.psup_tipe.currentText()
        sku_kode = self.psup_sku.currentData()
        sku_text = self.psup_sku.currentText()
        qty, harga = self.psup_qty.value(), self.psup_harga.value()
        
        if not sku_text or sku_text.startswith("--") or harga <= 0:
            return QMessageBox.warning(self, "Error", "SKU tidak terdaftar di Master Data!")
            
        self.cart_pengsup.append({
            "tipe": tipe, "sku_kode": sku_kode if sku_kode else sku_text,
            "nama_garapan": sku_text, "qty": qty, "harga": harga, "total": qty * harga
        })
        self.refresh_pengsup_table(); self.psup_qty.setValue(0)

    def del_pengsup_item(self):
        selected = self.table_psup.selectedItems()
        if not selected: return
        rows = sorted(list(set([item.row() for item in selected])), reverse=True)
        for r in rows: self.cart_pengsup.pop(r)
        self.refresh_pengsup_table()

    def refresh_pengsup_table(self):
        self.table_psup.setRowCount(0)
        total_pemasukan = 0
        
        for i, item in enumerate(self.cart_pengsup):
            self.table_psup.insertRow(i)
            self.table_psup.setItem(i, 0, QTableWidgetItem(item['tipe']))
            self.table_psup.setItem(i, 1, QTableWidgetItem(item['nama_garapan']))
            self.table_psup.setItem(i, 2, QTableWidgetItem(f"{item['qty']:g}"))
            self.table_psup.setItem(i, 3, QTableWidgetItem(f"Rp {item['harga']:,.0f}"))
            self.table_psup.setItem(i, 4, QTableWidgetItem(f"Rp {item['total']:,.0f}"))
            total_pemasukan += item['total']
            
        kain_mentah_tot = self.psup_kain_qty.value() * self.psup_kain_harga.value()
        gaji_kotor = total_pemasukan - kain_mentah_tot
        potong_bon = self.psup_potong_bon.value()
        gaji_bersih = gaji_kotor - potong_bon
        
        self.lbl_pemasukan.setText(f"Total Pemasukan Barang/Potongan: Rp {total_pemasukan:,.0f}")
        self.lbl_pengurang.setText(f"- Potongan Harga Kain Mentah: Rp {kain_mentah_tot:,.0f}")
        self.lbl_gaji_kotor_psup.setText(f"TOTAL DITERIMA (SBLM KASBON): Rp {gaji_kotor:,.0f}")
        self.lbl_gaji_bersih_psup.setText(f"GAJI BERSIH (NETT): Rp {gaji_bersih:,.0f}")

    def submit_pengsup(self):
        person_id = self.psup_person.currentData()
        if not person_id or not self.cart_pengsup: 
            return QMessageBox.warning(self, "Error", "Pilih pengsup & pastikan daftar garapan tidak kosong!")

        # 1. AMBIL NILAI KAIN MENTAH DARI UI ANDA
        kain_qty = self.psup_kain_qty.value()
        kain_harga = self.psup_kain_harga.value()
        kain_total = kain_qty * kain_harga

        # 2. HITUNG BARANG JADI VS JASA POTONGAN
        total_pemasukan_barang = 0.0
        total_potongan_jasa = 0.0

        for item in self.cart_pengsup:
            tipe_garapan = item.get('tipe', '')
            if "Potongan" in tipe_garapan:
                total_potongan_jasa += item['total'] # Jasa Potong
            else:
                total_pemasukan_barang += item['total'] # Barang Jadi

        # 3. RUMUS GRAND TOTAL SESUAI GAMBAR ANDA:
        gaji_kotor = total_pemasukan_barang - kain_total + total_potongan_jasa

        bon_lama = self.psup_bon_lama.value()
        tambah_bon = self.psup_tambah_bon.value()
        potong_bon = self.psup_potong_bon.value()
        gaji_bersih = gaji_kotor - potong_bon
        sisa_bon_akhir = bon_lama + tambah_bon - potong_bon

        try:
            tanggal_str = self.psup_date.date().toString("yyyy-MM-dd")
            run = SalaryRun(
                tipe="PENGSUP", person_id=person_id, tanggal_proses=tanggal_str, 
                gaji_kotor=gaji_kotor, bon_lama=bon_lama, tambah_bon=tambah_bon, 
                potong_bon=potong_bon, gaji_bersih=gaji_bersih, sisa_bon_akhir=sisa_bon_akhir
            )
            self.db.add(run)
            self.db.flush()

            # --- 4. SIMPAN KAIN MENTAH SEBAGAI ITEM RAHASIA KE SQLITE ---
            if kain_total > 0:
                line_mentah = SalaryLineItem(
                    salary_run_id=run.id, sku_id=None,
                    model_code="[KAIN_MENTAH]", # TAG Rahasia
                    qty=kain_qty, tarif_per_pcs=kain_harga, subtotal=-kain_total
                )
                self.db.add(line_mentah)

            # --- 5. SIMPAN TABEL GRID ---
            for item in self.cart_pengsup:
                tipe_garapan = item.get('tipe', '')
                prefix = "[POTONG] " if "Potongan" in tipe_garapan else "[BARANG] "
                
                line = SalaryLineItem(
                    salary_run_id=run.id, sku_id=item.get('sku_id'),
                    model_code=prefix + item['nama_garapan'], 
                    qty=item['qty'], tarif_per_pcs=item['harga'], subtotal=item['total']
                )
                self.db.add(line)

            # Update Kasbon
            balance = self.db.query(BonBalance).filter(BonBalance.person_id == person_id).first()
            if not balance:
                balance = BonBalance(person_id=person_id, saldo=0)
                self.db.add(balance)

            if tambah_bon > 0:
                self.db.add(BonMovement(person_id=person_id, tanggal=tanggal_str, tipe="TAMBAH", nominal=tambah_bon, sumber="PAYROLL_PENGSUP"))
                balance.saldo += tambah_bon
            if potong_bon > 0:
                self.db.add(BonMovement(person_id=person_id, tanggal=tanggal_str, tipe="POTONG_GAJI", nominal=potong_bon, sumber="PAYROLL_PENGSUP"))
                balance.saldo -= potong_bon

            self.db.commit()
            if hasattr(self, 'notifier') and self.notifier:
                print("[*] Broadcasting database changes to all menus...")
                self.notifier.database_changed.emit()
                
            # Trigger PDF
            from utils.pdf_engine import generate_salary_slip
            pdf_path = generate_salary_slip(run.id)
            if pdf_path and os.path.exists(pdf_path): os.startfile(pdf_path)

            QMessageBox.information(self, "Sukses", "Data berhasil disimpan dan PDF dibuat!")
            self.cart_pengsup.clear(); self.refresh_pengsup_table()
            self.psup_kain_qty.setValue(0); self.psup_kain_harga.setValue(0)
            self.psup_tambah_bon.setValue(0); self.psup_potong_bon.setValue(0)

        except Exception as e:
            self.db.rollback()
            self.db.expire_all()
            QMessageBox.critical(self, "Error", f"Gagal Database: {e}")

    def import_excel_pengsup(self):
        """
        Smart Import Draft Pengsup:
        Kebal terhadap perubahan nama sheet dan pergeseran baris metadata (Autopilot Scanning).
        """
        file_path, _ = QFileDialog.getOpenFileName(self, "Import Setoran Pengsup", "", "Excel Files (*.xlsx *.xls)")
        if not file_path: return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # [SOLUSI AMAN]: Cek jika sheet Daftar_Pemasukan ada, jika tidak ada pakai sheet aktif pertama
            if "Daftar_Pemasukan" in wb.sheetnames:
                ws = wb["Daftar_Pemasukan"]
            else:
                ws = wb.active # Terhindar dari error "File Salah" akibat nama sheet berubah
                
            # Inisialisasi nilai penampung awal
            kain_qty = 0.0
            kain_harga = 0.0
            tambah_bon = 0.0
            potong_bon = 0.0
            data_start_row = 8 # Fallback default posisi daftar barang

            # ========================================================
            # LOGIKA SCANNING SPASIAL (Mencari koordinat teks metadata)
            # ========================================================
            # Kita scan 20 baris teratas dan 5 kolom pertama untuk mencari keyword
            for r in range(1, min(20, ws.max_row + 1)):
                for c in range(1, min(5, ws.max_column + 1)):
                    cell_val = str(ws.cell(r, c).value or "").strip().lower()
                    
                    if not cell_val: continue
                    
                    # Bersihkan dan ambil nilai angka di sebelah kanan cell keyword (c + 1)
                    if "kain_qty" in cell_val or "kain qty" in cell_val:
                        val_raw = str(ws.cell(r, c+1).value or "0").replace(',', '')
                        kain_qty = float(val_raw) if val_raw else 0.0
                    elif "kain_harga" in cell_val or "harga kain" in cell_val:
                        val_raw = str(ws.cell(r, c+1).value or "0").replace('Rp', '').replace(',', '').strip()
                        kain_harga = float(val_raw) if val_raw else 0.0
                    elif "bon_tambah" in cell_val or "tambah bon" in cell_val:
                        val_raw = str(ws.cell(r, c+1).value or "0").replace('Rp', '').replace(',', '').strip()
                        tambah_bon = float(val_raw) if val_raw else 0.0
                    elif "bon_potong" in cell_val or "potong bon" in cell_val:
                        val_raw = str(ws.cell(r, c+1).value or "0").replace('Rp', '').replace(',', '').strip()
                        potong_bon = float(val_raw) if val_raw else 0.0
                    elif "sku" in cell_val or "nama barang" in cell_val or "jenis garapan" in cell_val:
                        data_start_row = r # Catat di baris mana tabel barang jadi dimulai

            # Masukkan hasil temuan angka ke dalam widget UI Anda secara aman
            self.psup_kain_qty.setValue(kain_qty)
            self.psup_kain_harga.setValue(kain_harga)
            self.psup_tambah_bon.setValue(tambah_bon)
            self.psup_potong_bon.setValue(potong_bon)
                
            # ========================================================
            # MEMBACA DAFTAR BARANG YANG BERADA DI BAWAH HEADER TABEL
            # ========================================================
            self.cart_pengsup.clear()
            for r in range(data_start_row + 1, ws.max_row + 1):
                if r % 10 == 0:
                    QApplication.processEvents()
                nama = ws.cell(r, 1).value
                # Lewati jika baris kosong atau berisi teks ringkasan "TOTAL"
                if not nama or str(nama).strip() == "" or "total" in str(nama).lower(): 
                    continue
                
                qty_val = str(ws.cell(r, 2).value or "0").replace(',', '').strip()
                harga_val = str(ws.cell(r, 3).value or "0").replace('Rp ', '').replace('.', '').replace(',', '').strip()
                tipe = str(ws.cell(r, 4).value or "Setor Barang Jadi (Kain)").strip()
                
                qty = float(qty_val) if qty_val else 0.0
                harga = float(harga_val) if harga_val else 0.0
                
                self.cart_pengsup.append({
                    "tipe": tipe, 
                    "sku_kode": str(nama), 
                    "nama_garapan": str(nama), 
                    "qty": qty, 
                    "harga": harga, 
                    "total": qty * harga
                })
                
            # Segarkan tampilan grid tabel di aplikasi Anda
            self.refresh_pengsup_table()
            QMessageBox.information(self, "Sukses", "Draf data Pengsup dan rincian Kain Mentah berhasil dimuat secara pintar tanpa halangan!")

        except Exception as e:
            QMessageBox.critical(self, "Error Import", f"Gagal membaca atau membedah struktur file Excel:\n{e}")

    # ==========================================
    # LOGIC BARU: GAJI PASUKAN (KARYAWAN)
    # ==========================================
    def load_karyawan_data(self):
        """Memuat semua nama karyawan ke dalam tabel grid pintar."""
        self.table_pasukan.setRowCount(0)
        karyawans = self.db.query(Person).filter(Person.person_type == 'KARYAWAN').order_by(Person.nama).all()

        for i, k in enumerate(karyawans):
            self.table_pasukan.insertRow(i)
            self.table_pasukan.setItem(i, 0, QTableWidgetItem(str(k.id)))

            item_nama = QTableWidgetItem(k.nama)
            item_nama.setFlags(item_nama.flags() & ~Qt.ItemFlag.ItemIsEditable) # Kunci kolom nama
            self.table_pasukan.setItem(i, 1, item_nama)

            # Kolom Input: Default 0
            for col in [2, 3, 4]:
                item_input = QTableWidgetItem("0")
                item_input.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table_pasukan.setItem(i, col, item_input)

            # Kolom Otomatis: Gaji Kotor
            item_kotor = QTableWidgetItem("Rp 0")
            item_kotor.setFlags(item_kotor.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table_pasukan.setItem(i, 5, item_kotor)

            # Kolom Info Kasbon
            balance = self.db.query(BonBalance).filter(BonBalance.person_id == k.id).first()
            bon_lama = balance.saldo if balance else 0
            item_bon = QTableWidgetItem(f"Rp {bon_lama:,.0f}")
            item_bon.setForeground(Qt.GlobalColor.magenta)
            item_bon.setFlags(item_bon.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table_pasukan.setItem(i, 6, item_bon)

            # Kolom Input: Potong Bon
            item_potong = QTableWidgetItem("0")
            item_potong.setForeground(Qt.GlobalColor.red)
            item_potong.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table_pasukan.setItem(i, 7, item_potong)

            # Kolom Otomatis: Gaji Bersih
            item_bersih = QTableWidgetItem("Rp 0")
            item_bersih.setForeground(Qt.GlobalColor.cyan)
            item_bersih.setFlags(item_bersih.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table_pasukan.setItem(i, 8, item_bersih)

    def import_absensi_excel(self):
        """Import Excel Mesin Fingerprint + Scanning Spasial + Autopilot Kasbon & Tarif Dinamis"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Pilih Excel Absensi Karyawan", "", "Excel Files (*.xlsx *.xls)")
        if not file_path: return

        try:
            self.load_karyawan_data()
            self.cache_absensi_harian = {} # 1. Reset Cache Harian untuk PDF
            
            self.lbl_file_absen.setText(f"File aktif: {os.path.basename(file_path)}")
            excel_data = pd.read_excel(file_path, sheet_name=None, header=None)
            rekap_dict = {}

            # =========================================================
            # LOGIKA SCANNING SPASIAL (MURNI MENGGUNAKAN KODE ANDA)
            # =========================================================
            for sheet_name, df in excel_data.items():
                if "Analisa Kehadiran" in str(sheet_name) or "Pengaturan Shift" in str(sheet_name):
                    continue

                # Cek lokasi tulisan "user id" sebagai jangkar (anchor) pembacaan
                uid_loc = [(r, c) for r in range(len(df)) for c in range(len(df.columns)) if "user id" in str(df.iloc[r, c]).strip().lower()]
                if not uid_loc: continue

                for r_uid, c_uid in uid_loc:
                    emp_id = ""
                    name_str = "Unknown"
                    
                    # Cari ID Karyawan
                    for offset in range(1, 10):
                        if c_uid + offset < len(df.columns) and str(df.iloc[r_uid, c_uid + offset]).strip() not in ['', 'nan', 'None']:
                            emp_id = str(df.iloc[r_uid, c_uid + offset]).strip().replace(".0", "")
                            break

                    # Cari Nama Karyawan
                    for r_nama in range(max(0, r_uid-2), r_uid+2):
                        for c_nama in range(max(0, c_uid-5), c_uid+5):
                            if c_nama < len(df.columns) and "nama" in str(df.iloc[r_nama, c_nama]).strip().lower():
                                for offset in range(1, 10):
                                    if c_nama + offset < len(df.columns) and str(df.iloc[r_nama, c_nama + offset]).strip() not in ['', 'nan', 'None']:
                                        name_str = str(df.iloc[r_nama, c_nama + offset]).strip().lower()
                                        break
                                break

                    # Inisialisasi Cache Rincian PDF untuk Karyawan ini
                    if name_str not in self.cache_absensi_harian:
                        self.cache_absensi_harian[name_str] = []

                    # Cari kolom Tanggal ("30 Sab", dsb)
                    date_col = -1
                    for c_test in range(c_uid, -1, -1):
                        for r_test in range(r_uid + 5, min(r_uid + 20, len(df))):
                            val = str(df.iloc[r_test, c_test]).strip()
                            if len(val) >= 4 and val[:2].isdigit() and " " in val:
                                date_col = c_test
                                break
                        if date_col != -1: break

                    if date_col == -1: continue

                    total_normal = 0
                    total_lembur = 0
                    hari_hadir = 0

                    for r_data in range(r_uid + 5, min(r_uid + 45, len(df))):
                        date_val = str(df.iloc[r_data, date_col]).strip()
                        if not (len(date_val) >= 4 and date_val[:2].isdigit()): continue

                        # Tarik dan Parse semua jam tap di hari tersebut
                        waktu_taps = [self.parse_waktu(df.iloc[r_data, c]) for c in range(date_col + 1, min(date_col + 15, len(df.columns))) if self.parse_waktu(df.iloc[r_data, c]) is not None]

                        if waktu_taps:
                            # Logika Hitungan Anda (Mencari tap masuk dan tap keluar)
                            min_t = round(min(waktu_taps) * 1440)
                            max_t = round(max(waktu_taps) * 1440)
                            
                            # Ubah angka desimal menjadi string "HH:MM"
                            jam_masuk_str = f"{int(min_t // 60):02d}:{int(min_t % 60):02d}"
                            
                            if len(waktu_taps) > 1:
                                jam_keluar_str = f"{int(max_t // 60):02d}:{int(max_t % 60):02d}"
                                diff = max_t - min_t
                                if diff < 0: diff += 1440
                                total_mnt = diff
                                hari_hadir += 1
                            else:
                                jam_keluar_str = "Lupa"
                                total_mnt = 0 
                                hari_hadir += 1
                            
                            menit_normal = min(total_mnt, 480)
                            lembur = max(0, total_mnt - 480)
                            
                            total_normal += menit_normal
                            total_lembur += lembur
                            
                            # ========================================================
                            # 2. SIMPAN CACHE PDF (DI DALAM LOOP HARIAN)
                            # ========================================================
                            self.cache_absensi_harian[name_str].append({
                                "tanggal": date_val[:6],  # Contoh Output: "30 Sab"
                                "masuk": jam_masuk_str,   # Contoh Output: "08:15"
                                "keluar": jam_keluar_str, # Contoh Output: "16:45" atau "Lupa"
                                "menit_normal": menit_normal,
                                "menit_lembur": lembur
                            })
                            
                    if name_str not in rekap_dict: rekap_dict[name_str] = {'hadir': 0, 'normal': 0, 'lembur': 0}
                    rekap_dict[name_str]['hadir'] += hari_hadir
                    rekap_dict[name_str]['normal'] += total_normal
                    rekap_dict[name_str]['lembur'] += total_lembur

            # ========================================================
            # 3. SEBAR DATA REKAP MINGGUAN KE GRID TABEL PYSIDE6
            # ========================================================
            all_balances = self.db.query(BonBalance).all()
            dict_balances = {b.person_id: b.saldo for b in all_balances}
            self.table_pasukan.blockSignals(True)
            
            from PySide6.QtWidgets import QApplication
            
            for r in range(self.table_pasukan.rowCount()):
                if r % 5 == 0: QApplication.processEvents()
                
                item_nama = self.table_pasukan.item(r, 1)
                if not item_nama: continue
                nama_tabel = item_nama.text().strip().lower()
                
                matched_name = None
                for ex_name in rekap_dict.keys():
                    if ex_name in nama_tabel or nama_tabel in ex_name:
                        matched_name = ex_name; break
                
                if matched_name:
                    data_karyawan = rekap_dict[matched_name]
                    val_hadir = str(data_karyawan['hadir'])
                    val_nrml = float(data_karyawan['normal'])
                    val_lmbr = float(data_karyawan['lembur'])

                    p_id = int(self.table_pasukan.item(r, 0).text())
                    bon_lama = dict_balances.get(p_id, 0.0)

                    trf_normal = self.tarif_normal.value()
                    trf_lembur = self.tarif_lembur.value()

                    gaji_kotor = (val_nrml * trf_normal) + (val_lmbr * trf_lembur)
                    potong_bon = bon_lama if gaji_kotor >= bon_lama else gaji_kotor
                    gaji_bersih = gaji_kotor - potong_bon

                    for col_idx in range(11):
                        if not self.table_pasukan.item(r, col_idx): 
                            self.table_pasukan.setItem(r, col_idx, QTableWidgetItem())
                        
                        item = self.table_pasukan.item(r, col_idx)
                        if col_idx in [3, 4, 5, 6, 9]:
                            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                        else:
                            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)

                    self.table_pasukan.item(r, 2).setText(val_hadir)
                    self.table_pasukan.item(r, 3).setText(f"{val_nrml:g}")
                    self.table_pasukan.item(r, 4).setText(f"{trf_normal:g}") 
                    self.table_pasukan.item(r, 5).setText(f"{val_lmbr:g}")
                    self.table_pasukan.item(r, 6).setText(f"{trf_lembur:g}") 
                    self.table_pasukan.item(r, 7).setText(f"Rp {gaji_kotor:,.0f}")
                    self.table_pasukan.item(r, 8).setText(f"Rp {bon_lama:,.0f}")
                    self.table_pasukan.item(r, 9).setText(f"{potong_bon:g}")
                    self.table_pasukan.item(r, 10).setText(f"Rp {gaji_bersih:,.0f}")

                    if gaji_bersih == 0 and gaji_kotor > 0:
                        for col_idx in range(11):
                            self.table_pasukan.item(r, col_idx).setBackground(QColor("#2d1f1f"))
                            self.table_pasukan.item(r, col_idx).setForeground(QColor(Theme.TEXT_MUTED))

            self.table_pasukan.blockSignals(False)
            self.recalc_pasukan()
            QMessageBox.information(self, "Berhasil", "Data Absensi berhasil dimuat! Ketuk ganda pada kolom Menit/Tarif/Potongan jika ingin melakukan revisi.")

        except Exception as e:
            self.table_pasukan.blockSignals(False)
            QMessageBox.critical(self, "Error", f"Gagal memproses file absensi:\n{e}")

    def recalc_pasukan(self):
        total_kotor, total_potong, total_bersih = 0.0, 0.0, 0.0
        for row in range(self.table_pasukan.rowCount()):
            try:
                it_kotor = self.table_pasukan.item(row, 7)
                it_potong = self.table_pasukan.item(row, 9)
                it_bersih = self.table_pasukan.item(row, 10)

                txt_kotor = it_kotor.text().replace('Rp ', '').replace(',', '').strip() if it_kotor and it_kotor.text() else "0"
                txt_potong = it_potong.text().replace('Rp ', '').replace(',', '').strip() if it_potong and it_potong.text() else "0"
                txt_bersih = it_bersih.text().replace('Rp ', '').replace(',', '').strip() if it_bersih and it_bersih.text() else "0"

                total_kotor += float(txt_kotor)
                total_potong += float(txt_potong)
                total_bersih += float(txt_bersih)
            except ValueError:
                continue

        if hasattr(self, 'lbl_grand_kotor'): self.lbl_grand_kotor.setText(f"Rp {total_kotor:,.0f}")
        if hasattr(self, 'lbl_grand_potong'): self.lbl_grand_potong.setText(f"Rp {total_potong:,.0f}")
        if hasattr(self, 'lbl_grand_bersih'): self.lbl_grand_bersih.setText(f"Rp {total_bersih:,.0f}")

    def submit_pasukan(self):
        tanggal_str = self.pasukan_date.date().toString("yyyy-MM-dd")
        run_ids_to_print = []
        jml_berhasil = 0
        try:
            person_ids_di_layar = []
            for row in range(self.table_pasukan.rowCount()):
                it_id = self.table_pasukan.item(row, 0)
                if it_id and it_id.text().strip():
                    person_ids_di_layar.append(int(it_id.text().strip()))
            balances_db = self.db.query(BonBalance).filter(BonBalance.person_id.in_(person_ids_di_layar)).all()
            dict_obj_balances = {b.person_id: b for b in balances_db}

            for row in range(self.table_pasukan.rowCount()):
                try:
                    it_id = self.table_pasukan.item(row, 0)
                    if not it_id or not it_id.text().strip(): continue
                    p_id = int(it_id.text().strip())

                    txt_hadir = self.table_pasukan.item(row, 2).text().strip() if self.table_pasukan.item(row, 2) else "0"
                    txt_normal = self.table_pasukan.item(row, 3).text().strip() if self.table_pasukan.item(row, 3) else "0"
                    txt_lembur = self.table_pasukan.item(row, 5).text().strip() if self.table_pasukan.item(row, 5) else "0"
                    
                    # [+] AMBIL TARIF DINAMIS LANGSUNG DARI SEL TABEL LAYAR (MENGAKOMODASI EDITAN KASIR)
                    txt_trf_normal = self.table_pasukan.item(row, 4).text().strip() if self.table_pasukan.item(row, 4) else "0"
                    txt_trf_lembur = self.table_pasukan.item(row, 6).text().strip() if self.table_pasukan.item(row, 6) else "0"

                    mnt_normal = float(txt_normal.replace(',', '')) if txt_normal else 0.0
                    mnt_lembur = float(txt_lembur.replace(',', '')) if txt_lembur else 0.0
                    
                    # Konversi nilai tarif dinamis per baris karyawan
                    trf_normal_aktual = float(txt_trf_normal.replace(',', '')) if txt_trf_normal else 0.0
                    trf_lembur_aktual = float(txt_trf_lembur.replace(',', '')) if txt_trf_lembur else 0.0

                    if mnt_normal == 0 and mnt_lembur == 0:
                        continue

                    it_kotor = self.table_pasukan.item(row, 7)
                    it_bon = self.table_pasukan.item(row, 8)
                    it_potong = self.table_pasukan.item(row, 9)
                    it_bersih = self.table_pasukan.item(row, 10)

                    gaji_kotor = float(it_kotor.text().replace('Rp ', '').replace(',', '').strip()) if it_kotor and it_kotor.text() else 0.0
                    bon_lama = float(it_bon.text().replace('Rp ', '').replace(',', '').strip()) if it_bon and it_bon.text() else 0.0
                    potong_bon = float(it_potong.text().replace('Rp ', '').replace(',', '').strip()) if it_potong and it_potong.text() else 0.0
                    gaji_bersih = float(it_bersih.text().replace('Rp ', '').replace(',', '').strip()) if it_bersih and it_bersih.text() else 0.0

                    sisa_bon_akhir = bon_lama - potong_bon
                    
                    # Catat tarif yang digunakan ke dalam keterangan string sebagai cadangan visual
                    keterangan = f"Hadir: {txt_hadir} | Normal: {mnt_normal:g} (@{trf_normal_aktual:g}) | Lembur: {mnt_lembur:g} (@{trf_lembur_aktual:g})"

                    # 1. Simpan data induk run gaji
                    run = SalaryRun(
                        tipe="PASUKAN_KARYAWAN", person_id=p_id, tanggal_proses=tanggal_str,
                        gaji_kotor=gaji_kotor, bon_lama=bon_lama, tambah_bon=0,
                        potong_bon=potong_bon, gaji_bersih=gaji_bersih, sisa_bon_akhir=sisa_bon_akhir,
                        catatan=keterangan
                    )
                    self.db.add(run)
                    self.db.flush() 
                    run_ids_to_print.append(run.id)

                    # =======================================================
                    # [+] SIMPAN TARIF AKTUAL SEBAGAI RINCIAN ITEM DETAIL KE SQLITE
                    # =======================================================
                    # Menyimpan Rincian Gaji Normal dengan Tarif Aktualnya
                    if mnt_normal > 0:
                        line_normal = SalaryLineItem(
                            salary_run_id=run.id, sku_id=None,
                            model_code="[GAJI_NORMAL]", 
                            qty=mnt_normal, tarif_per_pcs=trf_normal_aktual, subtotal=(mnt_normal * trf_normal_aktual)
                        )
                        self.db.add(line_normal)
                        
                    # Menyimpan Rincian Gaji Lembur dengan Tarif Aktualnya
                    if mnt_lembur > 0:
                        line_lembur = SalaryLineItem(
                            salary_run_id=run.id, sku_id=None,
                            model_code="[GAJI_LEMBUR]", 
                            qty=mnt_lembur, tarif_per_pcs=trf_lembur_aktual, subtotal=(mnt_lembur * trf_lembur_aktual)
                        )
                        self.db.add(line_lembur)

                    # =======================================================
                    # [+] SIMPAN HISTORI ABSENSI HARIAN DARI EXCEL CACHE
                    # =======================================================
                    p_nama = self.table_pasukan.item(row, 1).text().strip().lower() 
                    if hasattr(self, 'cache_absensi_harian') and self.cache_absensi_harian:
                        for cache_name, daily_list in self.cache_absensi_harian.items():
                            if p_nama in cache_name or cache_name in p_nama:
                                from data.models.salary import AttendanceRecord
                                for harian in daily_list:
                                    rec = AttendanceRecord(
                                        salary_run_id=run.id, person_id=p_id,
                                        tanggal=harian['tanggal'], tap_masuk=harian['masuk'], tap_keluar=harian['keluar'],
                                        menit_normal=harian['menit_normal'], menit_lembur=harian['menit_lembur']
                                    )
                                    self.db.add(rec)
                                break 

                    # Update Saldo Kasbon Karyawan
                    if potong_bon > 0:
                        balance = dict_obj_balances.get(p_id)
                        if balance:
                            balance.saldo -= potong_bon
                            self.db.add(BonMovement(person_id=p_id, tanggal=tanggal_str, tipe="POTONG_GAJI", nominal=potong_bon, sumber="PAYROLL_KARYAWAN"))

                    jml_berhasil += 1
                except ValueError:
                    continue

            self.db.commit()
            if hasattr(self, 'notifier') and self.notifier:
                self.notifier.database_changed.emit()
                
            if run_ids_to_print:
                # Triger pencetakan massal PDF
                for s_id in run_ids_to_print:
                    generate_salary_slip(s_id)
                
                export_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "exports", "slips")
                if os.path.exists(export_dir): os.startfile(export_dir)
                QMessageBox.information(self, "Sukses!", f"{len(run_ids_to_print)} Slip Gaji Karyawan berhasil disimpan dengan tarif kustom dan PDF telah dibuat!")
                self.load_karyawan_data() 
                self.recalc_pasukan()
            else:
                QMessageBox.warning(self, "Kosong", "Tidak ada data gaji karyawan yang valid untuk diproses.")

        except Exception as e:
            self.db.rollback()
            self.db.expire_all()
            QMessageBox.critical(self, "Error", f"Gagal menyimpan data gaji: {e}")

    def _on_pdf_batch_success(self, jumlah_slip, export_dir):
        """Dipanggil otomatis saat PDFWorker selesai 100%"""
        
        # 1. Kembalikan tombol seperti semula
        if hasattr(self, 'btn_submit_pasukan'):
            self.btn_submit_pasukan.setEnabled(True)
            self.btn_submit_pasukan.setText("Simpan && Cetak All")

        # 2. Buka folder Windows Explorer jika ada
        if export_dir and os.path.exists(export_dir):
            os.startfile(export_dir)

        # 3. Tampilkan pesan sukses
        QMessageBox.information(self, "Sukses!", f"{jumlah_slip} Slip Gaji Karyawan berhasil disimpan dan PDF telah dibuat di latar belakang!")
        
        # 4. Reset tabel (Sekarang aman dilakukan di akhir)
        self.load_karyawan_data() 
        self.recalc_pasukan()

    def _on_pdf_batch_error(self, err_msg):
        """Dipanggil jika terjadi error saat mencetak PDF di background"""
        
        # Kembalikan tombol
        if hasattr(self, 'btn_submit_pasukan'):
            self.btn_submit_pasukan.setEnabled(True)
            self.btn_submit_pasukan.setText("Simpan && Cetak All")
            
        QMessageBox.warning(self, "PDF Error", f"Data berhasil disimpan ke database, TETAPI gagal mencetak PDF:\n{err_msg}")
        
        # Tetap reset tabel karena data sudah masuk database
        self.load_karyawan_data() 
        self.recalc_pasukan()
    
    # ==========================================
    # FITUR BACKUP DRAFT / STACK KE EXCEL
    # ==========================================
    def export_excel_penjahit(self):
        """Mengekspor draft/stack garapan penjahit yang ada di keranjang saat ini ke Excel."""
        if not self.cart_penjahit:
            return QMessageBox.warning(self, "Keranjang Kosong", "Tabel garapan masih kosong! Tambahkan garapan terlebih dahulu sebelum di-export.")

        file_path, _ = QFileDialog.getSaveFileName(self, "Backup Draft Penjahit", "Draft_Gaji_Penjahit.xlsx", "Excel Files (*.xlsx)")
        if not file_path: return
        
        try:
            data = []
            for item in self.cart_penjahit:
                data.append({
                    "SKU/Garapan": item['nama_garapan'],
                    "Qty": item['qty'],
                    "Harga/Tarif": item['harga']
                    # Total tidak perlu diexport karena saat di-import akan dikalkulasi ulang otomatis
                })
                
            df = pd.DataFrame(data)
            df.to_excel(file_path, index=False)
            QMessageBox.information(self, "Sukses", "Draft garapan berhasil diamankan ke Excel!\nJika nanti ada salah cetak PDF, Anda bisa meng-import file ini kembali.")
        except Exception as e:
            QMessageBox.critical(self, "Error Export", f"Gagal mengexport draft: {e}")

    def export_excel_pengsup(self):
        if not self.cart_pengsup: return QMessageBox.warning(self, "Keranjang Kosong", "Tabel garapan masih kosong!")
        file_path, _ = QFileDialog.getSaveFileName(self, "Backup Draft Pengsup", "Draft_Totalan_Pengsup.xlsx", "Excel Files (*.xlsx)")
        if not file_path: return
        
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Daftar_Pemasukan"
            ws1.append(["NAMA", self.psup_person.currentText()])
            ws1.append(["TANGGAL", self.psup_date.date().toString("yyyy-MM-dd")])
            
            # --- MENYIMPAN KAIN MENTAH KE EXCEL ---
            ws1.append(["KAIN_QTY", self.psup_kain_qty.value()]) 
            ws1.append(["KAIN_HARGA", self.psup_kain_harga.value()]) 
            
            ws1.append(["BON_TAMBAH", self.psup_tambah_bon.value()])
            ws1.append(["BON_POTONG", self.psup_potong_bon.value()])
            ws1.append([])
            ws1.append(["SKU/Nama", "Qty", "Harga/Tarif", "Tipe/Kategori"])
            
            for item in self.cart_pengsup:
                ws1.append([item['nama_garapan'], item['qty'], item['harga'], item.get('tipe', 'Setor Barang Jadi (Kain)')])
                
            wb.save(file_path)
            QMessageBox.information(self, "Sukses", "Draft garapan beserta data Kain Mentah berhasil diamankan ke Excel!")
        except Exception as e:
            QMessageBox.critical(self, "Error Export", f"Gagal mengexport draft: {e}")

    def export_excel_pasukan(self):
        """Mengekspor format file Excel absensi karyawan yang otomatis berisi daftar semua nama karyawan terdaftar"""
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Format Excel Karyawan", "Format_Absensi_Karyawan.xlsx", "Excel Files (*.xlsx)")
        if not file_path: return
        try:
            karyawans = self.db.query(Person).filter(Person.person_type == 'KARYAWAN').order_by(Person.nama).all()
            data = []
            for k in karyawans:
                data.append({
                    "Nama Karyawan": k.nama,
                    "Hadir": 0,
                    "Menit Normal": 0,
                    "Menit Lembur": 0
                })
            
            if not data:
                data = [{"Nama Karyawan": "Nama Karyawan Contoh", "Hadir": 0, "Menit Normal": 0, "Menit Lembur": 0}]
                
            df = pd.DataFrame(data)
            df.to_excel(file_path, index=False)
            QMessageBox.information(self, "Sukses", "Format Excel Absensi Karyawan berhasil diexport!")
        except Exception as e:
            QMessageBox.critical(self, "Error Export", f"Gagal mengexport file template absensi karyawan: {e}")
    
    def on_pasukan_cell_edited(self, row, column):
        """UX Reaktif: Menghitung ulang otomatis jika Menit, TARIF, atau Potongan di-edit kasir"""
        if column in [3, 4, 5, 6, 9]:
            self.table_pasukan.blockSignals(True)
            try:
                it_nrml = self.table_pasukan.item(row, 3)
                it_trf_nrml = self.table_pasukan.item(row, 4)
                it_lmbr = self.table_pasukan.item(row, 5)
                it_trf_lmbr = self.table_pasukan.item(row, 6)
                it_bon_lama = self.table_pasukan.item(row, 8)
                it_potong = self.table_pasukan.item(row, 9)

                # Parsing angka murni desimal dari sel tabel secara aman
                val_nrml = float(it_nrml.text().strip()) if it_nrml and it_nrml.text().strip() else 0.0
                trf_nrml = float(it_trf_nrml.text().replace('Rp ', '').replace(',', '').strip()) if it_trf_nrml and it_trf_nrml.text().strip() else self.tarif_normal.value()
                
                val_lmbr = float(it_lmbr.text().strip()) if it_lmbr and it_lmbr.text().strip() else 0.0
                trf_lmbr = float(it_trf_lmbr.text().replace('Rp ', '').replace(',', '').strip()) if it_trf_lmbr and it_trf_lmbr.text().strip() else self.tarif_lembur.value()
                
                txt_bon = it_bon_lama.text().replace('Rp ', '').replace(',', '').strip() if it_bon_lama else "0"
                bon_lama = float(txt_bon) if txt_bon else 0.0

                # KALKULASI DINAMIS BERDASARKAN TARIF YANG SEDANG DI-EDIT
                gaji_kotor = (val_nrml * trf_nrml) + (val_lmbr * trf_lmbr)
                
                if column == 9: # Jika yang diedit kolom potongan, ikuti kemauan kasir
                    txt_potong = it_potong.text().replace(',', '').strip() if it_potong else "0"
                    potong_bon = float(txt_potong) if txt_potong else 0.0
                else: # Jika yang diedit menit/tarif, jalankan rumus autopilot kasbon
                    potong_bon = bon_lama if gaji_kotor >= bon_lama else gaji_kotor
                
                if potong_bon > bon_lama: potong_bon = bon_lama
                if potong_bon > gaji_kotor: potong_bon = gaji_kotor

                gaji_bersih = gaji_kotor - potong_bon

                # --- PERBAIKAN BUG NoneType ---
                # Pastikan SELURUH widget item (11 kolom) di baris ini tidak ada yang None
                for c in range(11):
                    if not self.table_pasukan.item(row, c): 
                        self.table_pasukan.setItem(row, c, QTableWidgetItem())

                # Tulis kembali hasil hitungan barunya ke layar secara rapi
                self.table_pasukan.item(row, 7).setText(f"Rp {gaji_kotor:,.0f}")
                self.table_pasukan.item(row, 9).setText(f"{potong_bon:g}")
                self.table_pasukan.item(row, 10).setText(f"Rp {gaji_bersih:,.0f}")
                
                # Update Gaya Warna secara dinamis (Aman karena semua sel sudah diinisialisasi di atas)
                if gaji_bersih == 0 and gaji_kotor > 0:
                    for col_idx in range(11):
                        self.table_pasukan.item(row, col_idx).setBackground(QColor("#2d1f1f"))
                        self.table_pasukan.item(row, col_idx).setForeground(QColor(Theme.TEXT_MUTED))
                else:
                    for col_idx in range(11):
                        self.table_pasukan.item(row, col_idx).setData(Qt.ItemDataRole.BackgroundRole, None)
                        self.table_pasukan.item(row, col_idx).setData(Qt.ItemDataRole.ForegroundRole, None)

            except ValueError:
                pass
            self.table_pasukan.blockSignals(False)
            self.recalc_pasukan()
    
    def closeEvent(self, event):
        self.db.close()
        super().closeEvent(event)